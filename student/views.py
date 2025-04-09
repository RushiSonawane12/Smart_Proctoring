from django.shortcuts import render,redirect,reverse
from . import forms,models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, timedelta
from exam import models as QMODEL
from teacher import models as TMODEL
import cv2
import numpy as np
from django.http import JsonResponse
import os
import uuid
from datetime import datetime
import openpyxl

#for showing signup/login button for student
def studentclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'student/studentclick.html')

def student_signup_view(request):
    userForm=forms.StudentUserForm()
    studentForm=forms.StudentForm()
    mydict={'userForm':userForm,'studentForm':studentForm}
    if request.method=='POST':
        userForm=forms.StudentUserForm(request.POST)
        studentForm=forms.StudentForm(request.POST,request.FILES)
        if userForm.is_valid() and studentForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            student=studentForm.save(commit=False)
            student.user=user
            student.save()
            my_student_group = Group.objects.get_or_create(name='STUDENT')
            my_student_group[0].user_set.add(user)
        return HttpResponseRedirect('studentlogin')
    return render(request,'student/studentsignup.html',context=mydict)

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_dashboard_view(request):
    dict={
    
    'total_course':QMODEL.Course.objects.all().count(),
    'total_question':QMODEL.Question.objects.all().count(),
    }
    return render(request,'student/student_dashboard.html',context=dict)

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_exam_view(request):
    student = QMODEL.Student.objects.get(user_id=request.user.id)
    courses = QMODEL.Course.objects.all()

    courses_with_exam = []
    for course in courses:
        results = QMODEL.Result.objects.filter(exam=course, student=student)
        exam_submitted = results.exists()
        course_data = {'course': course, 'examSubmitted': exam_submitted}
        courses_with_exam.append(course_data)

    return render(request, 'student/student_exam.html', {'courses_with_exam': courses_with_exam})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def take_exam_view(request,pk):

    course=QMODEL.Course.objects.get(id=pk)
    total_questions=QMODEL.Question.objects.all().filter(course=course).count()
    questions=QMODEL.Question.objects.all().filter(course=course)
    total_marks=0
    for q in questions:
        total_marks=total_marks + q.marks
    
    return render(request,'student/take_exam.html',{'course':course,'total_questions':total_questions,'total_marks':total_marks})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def start_exam_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    questions=QMODEL.Question.objects.all().filter(course=course)
    if request.method=='POST':
        pass
    response= render(request,'student/start_exam.html',{'course':course,'questions':questions})
    response.set_cookie('course_id',course.id)
    return response


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def calculate_marks_view(request):
    if request.COOKIES.get('course_id') is not None:
        course_id = request.COOKIES.get('course_id')
        course=QMODEL.Course.objects.get(id=course_id)
        
        total_marks=0
        questions=QMODEL.Question.objects.all().filter(course=course)
        for i in range(len(questions)):
            
            selected_ans = request.COOKIES.get(str(i+1))
            actual_answer = questions[i].answer
            if selected_ans == actual_answer:
                total_marks = total_marks + questions[i].marks
        student = models.Student.objects.get(user_id=request.user.id)
        result = QMODEL.Result()
        result.marks=total_marks
        result.exam=course
        result.student=student
        result.save()

        return HttpResponseRedirect('view-result')



@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def view_result_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/view_result.html',{'courses':courses})
    
@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check_exam_submitted(request,pk):
    course = QMODEL.Course.objects.get(id=pk)
    student = models.Student.objects.get(user_id=request.user.id)
    results = QMODEL.Result.objects.filter(exam=course, student=student)
    
    if results.exists(): # If there are results
        exam_attempted = True
    else:
        exam_attempted = False
        
    return render(request, 'student/student_exam.html', {'exam_attempted': exam_attempted})

    


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check_marks_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    student = models.Student.objects.get(user_id=request.user.id)
    results= QMODEL.Result.objects.all().filter(exam=course).filter(student=student)
    return render(request,'student/check_marks.html',{'results':results})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_marks_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/student_marks.html',{'courses':courses})
  

def detect_faces(image):
    # Load pre-trained Haar Cascade Classifier for face detection
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Detect faces in the image
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Count the number of faces
    return len(faces)


def count_people(request):
    try:
        # Get the image file from the request
        image_file = request.FILES.get('image')

        # Check if image file exists
        if not image_file:
            return JsonResponse({'error': 'No image file provided'})

        # Read the image file
        nparr = np.frombuffer(image_file.read(), np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        # Process the image and count faces
        faces_count = detect_faces(image)

        # Return the faces count as JSON response
        return JsonResponse({'faces_count': faces_count})

    except Exception as e:
        return JsonResponse({'error': str(e)})
    
def save_image(request):
    if request.method == 'POST' and request.FILES.get('image'):
        image = request.FILES['image']
        
        # Generate a unique filename
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        random_string = str(uuid.uuid4()).replace("-", "")[:6]  # Generate a random string
        file_extension = os.path.splitext(image.name)[1]  # Get the file extension
        unique_filename = f"suspicious_image_{timestamp}_{random_string}{file_extension}"

        # Construct the path to save the image
        save_path = os.path.join('suspiciousImage/', unique_filename)
        
        with open(save_path, 'wb+') as destination:
            for chunk in image.chunks():
                destination.write(chunk)
        
        return JsonResponse({'message': 'Image saved successfully.', 'filename': unique_filename})
    else:
        return JsonResponse({'error': 'Invalid request.'}, status=400)


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check_test(request,pk):
    print(pk)
    return render(request,"student/studentchec.html",context={"pk":pk})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check(request,pk):

    p = request.POST.get('password')
    #print(prntexc("test_stuff.xlsx"))
    if excel_pass_check(pk,p) == True :
        course = QMODEL.Course.objects.get(id=pk)
        total_questions = QMODEL.Question.objects.all().filter(course=course).count()
        questions = QMODEL.Question.objects.all().filter(course=course)
        total_marks = 0
        for q in questions:
            total_marks = total_marks + q.marks

        return render(request, 'student/take_exam.html',
                      {'course': course, 'total_questions': total_questions, 'total_marks': total_marks})
    else:
        pass
        # return render(request, "student/studentchec.html", context={"pk": pk})


import openpyxl
import os
def excel_pass_check(id,pas):
    workbook = openpyxl.load_workbook("test_stuff.xlsx")
    sheet_obj = workbook.active

    max_rows = sheet_obj.max_row
    print(max_rows)
    row = 0
    # Loop will print all columns name
    for i in range(2, max_rows + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        if int(id) == int(cell_obj.value) :
            row =  i
            break
    print(row,"dsda")
    if row != 0  :
        cell_obj = sheet_obj.cell(row=row, column=2)
        print("sdgahd",cell_obj)

        if str(pas) == str(cell_obj.value) :
            return True
        else:
            return False
    else:
        return False

def prntexc(names):
    workbook = openpyxl.load_workbook(names)
    sheet_obj = workbook.active
    max_col = sheet_obj.max_column
    max_rows = sheet_obj.max_row

    # Loop will print all columns name
    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            print(cell_obj.value ,end=" ")
        print()

