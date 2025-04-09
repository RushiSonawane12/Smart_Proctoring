import random

from django.shortcuts import render,redirect,reverse
from . import forms,models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, timedelta
from exam import models as QMODEL
from student import models as SMODEL
from exam import forms as QFORM

from django.core.mail import send_mail
from django.conf import settings



#for showing signup/login button for teacher
def teacherclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'teacher/teacherclick.html')

def teacher_signup_view(request):
    userForm=forms.TeacherUserForm()
    teacherForm=forms.TeacherForm()
    mydict={'userForm':userForm,'teacherForm':teacherForm}
    if request.method=='POST':
        userForm=forms.TeacherUserForm(request.POST)
        teacherForm=forms.TeacherForm(request.POST,request.FILES)
        if userForm.is_valid() and teacherForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            teacher=teacherForm.save(commit=False)
            teacher.user=user
            teacher.save()
            my_teacher_group = Group.objects.get_or_create(name='TEACHER')
            my_teacher_group[0].user_set.add(user)
        return HttpResponseRedirect('teacherlogin')
    return render(request,'teacher/teachersignup.html',context=mydict)



def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_dashboard_view(request):
    dict={
    
    'total_course':QMODEL.Course.objects.all().count(),
    'total_question':QMODEL.Question.objects.all().count(),
    'total_student':SMODEL.Student.objects.all().count()
    }

    return render(request,'teacher/teacher_dashboard.html',context=dict)

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_exam_view(request):
    return render(request,'teacher/teacher_exam.html')


@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_add_exam_view(request):
    courseForm=QFORM.CourseForm()
    if request.method=='POST':
        courseForm=QFORM.CourseForm(request.POST)
        if courseForm.is_valid():        
            courseForm.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/teacher/teacher-view-exam')
    return render(request,'teacher/teacher_add_exam.html',{'courseForm':courseForm})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_exam_view(request):
    courses = QMODEL.Course.objects.all()
    return render(request,'teacher/teacher_view_exam.html',{'courses':courses})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def delete_exam_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    course.delete()
    return HttpResponseRedirect('/teacher/teacher-view-exam')

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_student_view(request):
    dict={
    'total_student':SMODEL.Student.objects.all().count(),
    }
    return render(request,'teacher/teacher-student.html',context=dict)
  

@login_required(login_url='adminlogin')
def teacher_question_view(request):
    return render(request,'teacher/teacher_question.html')

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_add_question_view(request):
    questionForm=QFORM.QuestionForm()
    if request.method=='POST':
        questionForm=QFORM.QuestionForm(request.POST)
        if questionForm.is_valid():


            course = QMODEL.Course.objects.get(id=request.POST.get('courseID'))
            qcnt=QMODEL.Question.objects.filter(course_id=course).count()
            check=0


            if(int(qcnt)<int(course.question_number)):
                check=1
                question=questionForm.save(commit=False)
                course=QMODEL.Course.objects.get(id=request.POST.get('courseID'))
                question.course=course
                question.save()
                pass_word_test = random.randint(1000, 9999)
                add_user_to_excel("test_stuff.xlsx", request.POST.get('courseID'), pass_word_test)




            qcnttest=QMODEL.Question.objects.filter(course_id=course).count()

            if(int(qcnttest)==int(course.question_number) and  check==1):
                lista = []
                students = SMODEL.Student.objects.all()
                for student in students:
                    # print(student.user.username)
                    lista.append(student.user.username)
                send_simple_email(lista, course, " " + "Password " + f"for the Test  = {str(pass_word_test)}")

            # pass_word_test = random.randint(1000,9999)
            try :
                pass
                # add_user_to_excel("test_stuff",request.POST.get('courseID'),pass_word_test)
            except :
                pass



            # print(course.course_name)
            # print(course.question_number)
            #
            #
            # print(QMODEL.Question.objects.filter(course_id=course).count())

            # QMODEL.Course.objects.all()


            # send_simple_email(lista,course,"test")


        else:
            print("form is invalid")
        return HttpResponseRedirect('/teacher/teacher-view-question')
    return render(request,'teacher/teacher_add_question.html',{'questionForm':questionForm})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_question_view(request):
    courses= QMODEL.Course.objects.all()
    return render(request,'teacher/teacher_view_question.html',{'courses':courses})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def see_question_view(request,pk):
    questions=QMODEL.Question.objects.all().filter(course_id=pk)
    return render(request,'teacher/see_question.html',{'questions':questions})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def remove_question_view(request,pk):
    question=QMODEL.Question.objects.get(id=pk)
    question.delete()
    return HttpResponseRedirect('/teacher/teacher-view-question')


# views.py or any other suitable location in your app


def send_simple_email(recipient_list,subject,message):
    subject = 'New test Added in ' + str(subject)
    message = 'new test added. ' + message
    from_email = settings.DEFAULT_FROM_EMAIL
    # recipient_list = ['patilap1854@gmail.com']


    # test= SMODEL.Student.objects.all()

    send_mail(subject, message, from_email, recipient_list)

# Call this function wherever needed in your code.


 # students=SMODEL.Student.objects.all()
 #
 #    lista=[]
 #    for student in students:
 #        print(student.user.username)
 #        lista.append(student.user.username)
 #
 #    send_simple_email(lista)

import openpyxl
import os


def add_user_to_excel(file_name, username, password):
    # Check if the file exists
    if not os.path.exists(file_name):
        # Create a new workbook and add headers if the file does not exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(["Username", "Password"])
        workbook.save(file_name)

    # Open the existing workbook
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook["Users"]

    # Add the username and password to the next available row
    sheet.append([username, password])

    # Save the workbook
    workbook.save(file_name)

def prntexc(names):
    workbook = openpyxl.load_workbook(names)
    sheet_obj = workbook.active
    max_col = sheet_obj.max_column
    max_rows = sheet_obj.max_row

    # Loop will print all columns name
    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            print(cell_obj.value ,end=" ")
        print()


